import sqlite3
import pandas as pd 
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import matplotlib.pyplot as plt
from openpyxl.drawing.image import Image as XLImage
import io

conn = sqlite3.connect('data/sales.db')
cursor = conn.cursor()

df = pd.read_sql_query("SELECT * FROM sales", conn)

print("Shape:", df.shape)
print("\nFirst 5 rows:")
print(df.head())

# Sales by region
region_summary = df.groupby('region')['total_amount'].agg(['sum', 'count']).round(2)
region_summary.columns = ['Total Revenue', 'Total Transactions']
region_summary = region_summary.sort_values('Total Revenue', ascending=False)

# Sales by product
product_summary = df.groupby('product')['total_amount'].agg(['sum', 'count']).round(2)
product_summary.columns = ['Total Revenue', 'Total Transactions']
product_summary = product_summary.sort_values('Total Revenue', ascending=False)

# Sales by payment_method
payment_method_summary = df.groupby('payment_method')['total_amount'].agg(['sum', 'count']).round(2)
payment_method_summary.columns = ['Total Revenue', 'Total Transactions']
payment_method_summary = payment_method_summary.sort_values('Total Revenue', ascending=False)

output_path = 'output/sales_report.xlsx'

with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
    df.to_excel(writer, sheet_name='Raw Data', index=False)
    region_summary.to_excel(writer, sheet_name='Sales by Region')
    product_summary.to_excel(writer, sheet_name='Sales by Product')
    payment_method_summary.to_excel(writer, sheet_name='Sales by Payment Method')

print("Basic report written. Now applying styles...")
conn.close()

wb = load_workbook(output_path)


# Colors
DARK_BLUE  = '1F3864'
MID_BLUE   = '2E75B6'
LIGHT_GREY = 'F2F2F2'
WHITE      = 'FFFFFF'

# Fonts
header_font = Font(name='Arial', bold=True, color=WHITE, size=11)
data_font   = Font(name='Arial', size=10)

# Fills
header_fill = PatternFill('solid', start_color=MID_BLUE)
alt_fill    = PatternFill('solid', start_color=LIGHT_GREY)

# Alignment
center = Alignment(horizontal='center', vertical='center')
left   = Alignment(horizontal='left', vertical='center')

# Border
thin   = Side(style='thin', color='CCCCCC')
border = Border(left=thin, right=thin, top=thin, bottom=thin)



# Styling the header row
def style_sheet(ws):
    # Style header row
    for cell in ws[1]:
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = center
        cell.border    = border

    # Style data rows with alternating colors
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        fill = alt_fill if row_idx % 2 == 0 else PatternFill()
        for cell in row:
            cell.font      = data_font
            cell.alignment = center
            cell.border    = border
            cell.fill      = fill
            if cell.column == 2:
                cell.number_format = '$#,##0.00'

    # Column widths
    for col in ws.columns:
        max_length = max((len(str(cell.value)) if cell.value else 0) for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max_length + 6

#style sheets
style_sheet(wb['Sales by Region'])
style_sheet(wb['Sales by Product'])
style_sheet(wb['Sales by Payment Method'])

def add_matplotlib_chart(ws, data_series, labels, title, position):
    fig, ax = plt.subplots(figsize=(8, 5))
    bars = ax.bar(labels, data_series, color='#2E75B6', edgecolor='white')
    ax.set_title(title, fontsize=14, fontweight='bold', pad=15)
    ax.set_xlabel(ax.get_xlabel(), fontsize=11)
    ax.set_ylabel('Revenue ($)', fontsize=11)
    ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, _: f'${x:,.0f}'))
    plt.xticks(rotation=15, ha='right')
    plt.tight_layout()

    img_buffer = io.BytesIO()
    plt.savefig(img_buffer, format='png', dpi=150)
    img_buffer.seek(0)
    plt.close()

    img = XLImage(img_buffer)
    ws.add_image(img, position)

# Add charts
def add_matplotlib_chart(ws, data_series, labels, title, xlabel, position):
    fig, ax = plt.subplots(figsize=(7, 4))
    colors = ['#2E75B6', '#1F3864', '#70AD47', '#ED7D31', '#FFC000',
              '#FF0000', '#9DC3E6', '#A9D18E', '#F4B183', '#C55A11']
    ax.bar(labels, data_series, color=colors[:len(labels)], edgecolor='white')
    ax.set_title(title, fontsize=13, fontweight='bold', pad=12)
    ax.set_xlabel(xlabel, fontsize=10, labelpad=10)
    ax.set_ylabel('Revenue ($)', fontsize=10, labelpad=10)
    ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, _: f'${x:,.0f}'))
    plt.xticks(rotation=15, ha='right', fontsize=9)
    plt.tight_layout()

    img_buffer = io.BytesIO()
    plt.savefig(img_buffer, format='png', dpi=150, bbox_inches='tight')
    img_buffer.seek(0)
    plt.close()

    img = XLImage(img_buffer)
    ws.add_image(img, position)

add_matplotlib_chart(wb['Sales by Region'],
    region_summary['Total Revenue'].values,
    region_summary.index.tolist(),
    'Revenue by Region', 'Region', 'E2')

add_matplotlib_chart(wb['Sales by Product'],
    product_summary['Total Revenue'].values,
    product_summary.index.tolist(),
    'Revenue by Product', 'Product', 'E2')

add_matplotlib_chart(wb['Sales by Payment Method'],
    payment_method_summary['Total Revenue'].values,
    payment_method_summary.index.tolist(),
    'Revenue by Payment Method', 'Payment Method', 'E2')

# Style Raw Data sheet
ws_raw = wb['Raw Data']

# Style header row
for cell in ws_raw[1]:
    cell.font      = header_font
    cell.fill      = header_fill
    cell.alignment = center
    cell.border    = border

# Auto column widths
for col in ws_raw.columns:
    max_length = max((len(str(cell.value)) if cell.value else 0) for cell in col)
    ws_raw.column_dimensions[get_column_letter(col[0].column)].width = max_length + 4

# Freeze the header row so it stays visible when scrolling
ws_raw.freeze_panes = 'A2'

wb.save(output_path)